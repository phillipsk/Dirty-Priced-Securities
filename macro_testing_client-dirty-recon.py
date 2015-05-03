from openpyxl import Workbook, load_workbook

filename = 'sda_2015.xlsx'

wb1 = Workbook()
ws1 = wb1.active

ws1['G1'] = 'Path'
ws1.title = 'Main'

adf = "Dirty Securities 04222015.xlsx"
f = "F:\\ana\\xlmacro\\" + adf

wb2 = load_workbook(f)

ws2 = wb2.active

ws2['H1'] = 'Recon2'
ws2.title = 'Main2'

#print wb2.get_sheet_names()

wb1.save(filename)
wb2.save(f)