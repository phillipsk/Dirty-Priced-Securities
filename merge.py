#!/usr/bin/env python
# coding=utf8

import sys
from openpyxl.reader.excel import load_workbook

DAILY_VALUES_COLUMN = 1
RECONCILED_VALUES_COLUMN = 6
DELETED_ITEMS_COLUMN = 11

def copy_to_column(values, sheet, column, start=1):
    for i, v in enumerate(values, start):
        sheet.cell(column=column, row=i).value = v

def merge(yearly_file_name, daily_file_name):
    daily = load_workbook(daily_file_name)
    reconciled = load_workbook(yearly_file_name)
    reconciled_sheet = reconciled.worksheets[0]

    deleted_items = []

    daily_values_worksheet = daily.worksheets[0]

    daily_values = set(c.value for c in daily_values_worksheet.columns[DAILY_VALUES_COLUMN-1][1:])
    old_values = set(c.value for c in reconciled_sheet.columns[RECONCILED_VALUES_COLUMN-1][1:])
    reconciled_values = set()

    for v in old_values:
        if v in daily_values:
            reconciled_values.add(v)
        else:
            deleted_items.append(v)

    reconciled_values.update(daily_values)

    log_start_row = 1

    while reconciled_sheet.cell(column=DELETED_ITEMS_COLUMN, row=log_start_row).value:
        log_start_row += 1

    copy_to_column(deleted_items, reconciled_sheet, DELETED_ITEMS_COLUMN, log_start_row)
    copy_to_column(reconciled_values, reconciled_sheet, RECONCILED_VALUES_COLUMN, 2)
    copy_to_column([None]*len(deleted_items), reconciled_sheet, RECONCILED_VALUES_COLUMN,
                   2 + len(reconciled_values))

    reconciled.save(yearly_file_name)


if __name__ == "__main__":
    merge(sys.argv[1], sys.argv[2])
